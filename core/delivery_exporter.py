"""交付资源一键提取：按歌曲生命周期并行处理伴奏/MIDI/歌词/段落信息导出。

按歌曲工作流：每首歌固定按 伴奏合成 → MIDI处理 → 歌词处理 → 段落信息导出
的顺序串行执行勾选项目，歌曲之间并行（默认 3 首并发）；交付总表作为整体
步骤与歌曲循环并行，缓存表命中优先、未命中重新提取（不下载直链、不导歌词）。

同一首歌的各项目共享一次音频校准（见 core.audio_calibration 的校准缓存），
校准音频优先复用 JSON 原目录缓存，未命中才重新下载；歌词仅导出原文歌词。
段落信息逐曲收集后统一写入 Excel。单曲某项目失败不影响该曲其他项目与其他曲目。
"""
from __future__ import annotations

import json
import os
import shutil
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from core.audio_downloader import (
    CACHE_DIR_NAME,
    AudioDownloadOptions,
    export_song_audio,
)
from core.lyric_exporter import (
    collect_section_export_rows,
    export_song_lyrics,
    write_sections_excel,
)
from core.metadata_exporter import METADATA_EXCEL_NAME, export_songs_metadata
from core.midi_exporter import export_song
from core.parser import SongData, load_song_json

# 项目标识与显示名（PROJECT_NAMES 亦用于日志框展示）
PROJECT_AUDIO = "audio"
PROJECT_MIDI = "midi"
PROJECT_LYRIC = "lyric"
PROJECT_SECTIONS = "sections"
PROJECT_METADATA = "metadata"

PROJECT_NAMES: dict[str, str] = {
    PROJECT_AUDIO: "伴奏处理",
    PROJECT_MIDI: "MIDI处理",
    PROJECT_LYRIC: "歌词处理",
    PROJECT_SECTIONS: "段落信息导出",
    PROJECT_METADATA: "交付总表导出",
}

# 交付总表最终文件名（需求④）
DELIVERY_METADATA_NAME = "资源产出交付总表.xlsx"
# 输出目录子文件夹（需求③）
SUBDIR_AUDIO = "合成伴奏"
SUBDIR_LYRIC = "歌词处理"
SUBDIR_MIDI = "MIDI处理"

# openpyxl 不支持并发保存同一文件，串行化两个 Excel 写入项目
_EXCEL_LOCK = threading.Lock()

# 每首歌内部固定执行顺序：伴奏 → MIDI → 歌词 → 段落（交付总表为整体步骤）
_SONG_PROJECT_ORDER = (PROJECT_AUDIO, PROJECT_MIDI, PROJECT_LYRIC, PROJECT_SECTIONS)
# 结果展示顺序（与日志框一致）：伴奏 → MIDI → 歌词 → 段落 → 交付总表
_PROJECT_DISPLAY_ORDER = (
    PROJECT_AUDIO,
    PROJECT_MIDI,
    PROJECT_LYRIC,
    PROJECT_SECTIONS,
    PROJECT_METADATA,
)
DEFAULT_SONG_CONCURRENCY = 3


@dataclass
class DeliveryProjectResult:
    name: str  # PROJECT_NAMES 中的中文名
    success: list[str] = field(default_factory=list)  # 输出文件绝对路径
    failed: list[tuple[str, str]] = field(default_factory=list)  # (json路径, 原因)；整项目失败时路径为 ""
    notes: list[str] = field(default_factory=list)  # 缓存复用/校准等说明


@dataclass
class DeliveryExportResult:
    projects: list[DeliveryProjectResult] = field(default_factory=list)
    error: str | None = None  # 编排层整体异常（扫描/启动阶段）


class _ProgressAggregator:
    """各项目报告自身进度，聚合为加权平均单一进度。

    各项目处理同一批 JSON，曲目数相同，故各项目等权。
    """

    def __init__(self, total: int, callback: Callable[[float, str], None] | None):
        self._lock = threading.Lock()
        self._last: dict[str, float] = {}  # 每个项目最近一次完成比例 0..1
        self._done: dict[str, int] = {}  # 每个项目已完成的歌曲数
        self._total = total
        self._cb = callback

    def wrap(self, name: str) -> Callable[[float, str], None]:
        """返回按给定完成比例报告的进度回调（整体步骤如交付总表使用）。"""

        def report(fraction: float, message: str) -> None:
            with self._lock:
                self._last[name] = max(0.0, min(1.0, fraction))
                overall = sum(self._last.values()) / self._total
            if self._cb:
                self._cb(overall * 100.0, f"[{PROJECT_NAMES[name]}] {message}")

        return report

    def advance(self, name: str, total_songs: int, message: str) -> None:
        """报告该项目又完成一首歌（歌曲工作流每完成一个项目步骤调用一次）。"""
        with self._lock:
            done = self._done.get(name, 0) + 1
            self._done[name] = done
            self._last[name] = done / total_songs if total_songs else 1.0
            overall = sum(self._last.values()) / self._total
        if self._cb:
            self._cb(overall * 100.0, f"[{PROJECT_NAMES[name]}] {message}")


class _ProjectCollector:
    """线程安全地聚合单个项目的成功/失败/说明（歌曲 worker 并发写入）。"""

    def __init__(self, name: str):
        self.name = name
        self.success: list[str] = []
        self.failed: list[tuple[str, str]] = []
        self.notes: list[str] = []
        self._lock = threading.Lock()

    def add_success(self, paths: str | list[str]) -> None:
        if isinstance(paths, str):
            paths = [paths]
        with self._lock:
            self.success.extend(paths)

    def add_failure(self, path: str, reason: str) -> None:
        with self._lock:
            self.failed.append((path, reason))

    def add_notes(self, notes: list[str]) -> None:
        with self._lock:
            self.notes.extend(notes)

    def result(self) -> DeliveryProjectResult:
        with self._lock:
            return DeliveryProjectResult(
                name=self.name,
                success=list(self.success),
                failed=list(self.failed),
                notes=list(self.notes),
            )


def _run_audio_for_song(song: SongData, output_dir: str) -> str:
    """伴奏：按音频下载模块页面默认参数（merge_har_drum / original / wav / 44100，响度限幅分轨电平均开启）。"""
    options = AudioDownloadOptions(
        content="merge_har_drum",
        key_mode="original",
        output_format="wav",
        sample_rate=44100,
        # 默认立体声输出（与音频下载模块页面默认一致）
        channels="stereo",
        # 音频下载模块页面默认：响度 -12 LUFS、限幅 -1 dBTP、分轨电平 -6 dB 均勾选
        loudness_enabled=True,
        loudness_lufs=-12.0,
        limiter_enabled=True,
        limiter_db=-1.0,
        track_gain_enabled=True,
        track_gain_db=-6.0,
    )
    sub = Path(output_dir) / SUBDIR_AUDIO
    return export_song_audio(song, str(sub), options)


def _run_midi_for_song(song: SongData, output_dir: str) -> tuple[list[str], list[str]]:
    """MIDI：按 MIDI 导出模块默认参数（merge_same，校准开启），返回 (输出文件, 校准日志)。"""
    sub = Path(output_dir) / SUBDIR_MIDI
    calibration_log: list[str] = []
    exported = export_song(
        song,
        str(sub),
        "merge_same",
        write_tempo=False,
        write_lyrics=True,
        lyric_granularity="word",
        lower_octave=True,
        write_section_markers=False,
        exclude_rap_sections=True,
        remove_non_melody_notes=True,
        time_offset_ms=0,
        audio_reference_calibration=True,
        calibration_log=calibration_log,
    )
    return exported, calibration_log


def _run_lyric_for_song(song: SongData, output_dir: str) -> tuple[str, list[str]]:
    """歌词：按歌词导出模块默认参数（ksc-txt / all / 原歌词，校准开启）。"""
    sub = Path(output_dir) / SUBDIR_LYRIC
    calibration_log: list[str] = []
    out = export_song_lyrics(
        song,
        str(sub),
        lyric_format="ksc-txt",
        part="all",
        lyric_field="ori",
        title_lang="origin",
        artist_lang="origin",
        time_offset_ms=0,
        audio_reference_calibration=True,
        calibration_log=calibration_log,
    )
    return out, calibration_log


def _run_sections_for_song(song: SongData) -> list[tuple[str, str, str, str, str, str, str]]:
    """段落：逐曲收集段落表格行（统一写 Excel 由编排层完成）。"""
    return collect_section_export_rows(
        song,
        title_lang="origin",
        artist_lang="origin",
        time_offset_ms=0,
        audio_reference_calibration=True,
    )


def _process_song(
    path: str,
    output_dir: str,
    enabled: set[str],
    collectors: dict[str, _ProjectCollector],
    section_rows: list,
    section_rows_lock: threading.Lock,
    total_songs: int,
    aggregator: _ProgressAggregator,
) -> None:
    """处理一首歌：按 伴奏→MIDI→歌词→段落 固定顺序串行执行勾选项目。

    单项目失败记录到该项目并继续下一项目；JSON 解析失败则全部启用项目记录失败。
    """
    basename = os.path.basename(path)
    try:
        song = load_song_json(path, "ori")
    except Exception as exc:
        for key in _SONG_PROJECT_ORDER:
            if key in enabled:
                collectors[key].add_failure(path, f"JSON 解析失败: {exc}")
        return

    for key in _SONG_PROJECT_ORDER:
        if key not in enabled:
            continue
        try:
            if key == PROJECT_AUDIO:
                out = _run_audio_for_song(song, output_dir)
                collectors[key].add_success(out)
            elif key == PROJECT_MIDI:
                outs, calibration_log = _run_midi_for_song(song, output_dir)
                collectors[key].add_success(outs)
                if calibration_log:
                    collectors[key].add_notes(
                        [f"{basename}: {calibration_log[0]}"]
                    )
            elif key == PROJECT_LYRIC:
                out, calibration_log = _run_lyric_for_song(song, output_dir)
                collectors[key].add_success(out)
                if calibration_log:
                    collectors[key].add_notes(
                        [f"{basename}: {calibration_log[0]}"]
                    )
            elif key == PROJECT_SECTIONS:
                rows = _run_sections_for_song(song)
                with section_rows_lock:
                    section_rows.extend(rows)
        except Exception as exc:
            collectors[key].add_failure(path, str(exc))
        finally:
            aggregator.advance(key, total_songs, f"正在处理: {basename}")


def _collect_mr_ids(json_paths: list[str]) -> set[int]:
    """轻量提取输入 JSON 的 mr_id 集合（顶层字段，无需完整解析）。"""
    ids: set[int] = set()
    for path in json_paths:
        with open(path, "r", encoding="utf-8") as f:
            ids.add(int(json.load(f).get("mr_id", 0)))
    return ids


def _read_cached_metadata_msids(cache_table: Path) -> set[int]:
    """读取缓存元数据表第一列（MSID）集合，表头及非数字单元格跳过。"""
    from openpyxl import load_workbook

    wb = load_workbook(cache_table, read_only=True, data_only=True)
    try:
        ws = wb.active
        msids: set[int] = set()
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            value = row[0]
            if value is None:
                continue
            try:
                msids.add(int(value))
            except (TypeError, ValueError):
                continue
        return msids
    finally:
        wb.close()


def _find_reusable_cache_table(json_paths: list[str], mr_ids: set[int]) -> Path | None:
    """按 JSON 父目录去重查找缓存表，任一表 MSID 集合覆盖全部输入即复用。"""
    needed = set(mr_ids)
    seen_parents: set[str] = set()
    for path in json_paths:
        parent = str(Path(path).parent)
        if parent in seen_parents:
            continue
        seen_parents.add(parent)
        cache_table = Path(parent) / CACHE_DIR_NAME / METADATA_EXCEL_NAME
        if not cache_table.is_file():
            continue
        try:
            cached = _read_cached_metadata_msids(cache_table)
        except Exception:
            continue
        if needed <= cached:
            return cache_table
    return None


def _run_metadata_project(
    json_paths: list[str],
    output_dir: str,
    progress_cb: Callable[[float, str], None],
) -> DeliveryProjectResult:
    """交付总表：缓存表可复用则拷贝重命名，否则重新提取（不下载直链、不导歌词）。"""
    result = DeliveryProjectResult(name=PROJECT_NAMES[PROJECT_METADATA])
    mr_ids = _collect_mr_ids(json_paths)

    # 分支 A：缓存可复用 → 直接拷贝并重命名
    cache_table = _find_reusable_cache_table(json_paths, mr_ids)
    if cache_table is not None:
        dest = Path(output_dir) / DELIVERY_METADATA_NAME
        cached_count = len(_read_cached_metadata_msids(cache_table))
        with _EXCEL_LOCK:
            shutil.copy2(cache_table, dest)
        result.success.append(str(dest))
        result.notes.append(
            f"命中缓存复用: {cache_table}（缓存 {cached_count} 首，"
            f"覆盖全部输入 {len(mr_ids)} 首）"
        )
        return result

    # 分支 B：重新提取（仅输出元数据表格；内部自动更新各目录缓存副本）
    def on_progress(index: int, total: int, name: str):
        if name.startswith("重试:"):
            progress_cb(1.0, f"重试: {name}")
        else:
            progress_cb(index / total, f"正在处理: {name}")

    with _EXCEL_LOCK:
        extracted = export_songs_metadata(
            json_paths,
            output_dir,
            download_resources=False,
            export_lyrics=False,
            progress_callback=on_progress,
        )
    final_path = os.path.join(output_dir, DELIVERY_METADATA_NAME)
    os.replace(extracted.excel_path, final_path)
    result.success.append(final_path)
    result.failed.extend(extracted.failed)
    result.notes.append(
        f"缓存缺失或不完整，重新提取（{len(mr_ids)} 首，未下载直链资源）"
    )
    return result


def run_delivery_export(
    json_paths: list[str],
    output_dir: str,
    *,
    enabled: set[str],
    max_workers: int | None = None,
    progress_callback: Callable[[float, str], None] | None = None,
) -> DeliveryExportResult:
    """按歌曲生命周期并行执行勾选项目并输出到 output_dir。

    enabled 为 PROJECT_* 常量子集。每首歌固定按 伴奏→MIDI→歌词→段落 串行执行
    勾选的资源项目，歌曲之间并行（max_workers 为歌曲级并发数，默认 3）；
    交付总表作为整体步骤与歌曲循环并行（不占歌曲并发槽位）。
    单曲某项目失败不影响该曲其他项目与其他曲目。
    progress_callback(百分比 0-100, 状态文案) 由各项目进度加权聚合而来。
    """
    enabled = set(enabled)
    tasks = [key for key in _PROJECT_DISPLAY_ORDER if key in enabled]
    if not tasks:
        return DeliveryExportResult(projects=[])

    song_concurrency = max(1, max_workers or DEFAULT_SONG_CONCURRENCY)
    has_metadata = PROJECT_METADATA in enabled
    aggregator = _ProgressAggregator(len(tasks), progress_callback)
    collectors = {key: _ProjectCollector(PROJECT_NAMES[key]) for key in tasks}
    section_rows: list[tuple[str, str, str, str, str, str, str]] = []
    section_rows_lock = threading.Lock()

    worker_count = song_concurrency + (1 if has_metadata else 0)
    metadata_result: DeliveryProjectResult | None = None
    with ThreadPoolExecutor(
        max_workers=worker_count, thread_name_prefix="delivery"
    ) as pool:
        futures = {
            pool.submit(
                _process_song,
                path,
                output_dir,
                enabled,
                collectors,
                section_rows,
                section_rows_lock,
                len(json_paths),
                aggregator,
            ): "song"
            for path in json_paths
        }
        if has_metadata:
            futures[
                pool.submit(
                    _run_metadata_project,
                    json_paths,
                    output_dir,
                    aggregator.wrap(PROJECT_METADATA),
                )
            ] = PROJECT_METADATA

        for future in as_completed(futures):
            kind = futures[future]
            if kind == PROJECT_METADATA:
                # 单曲 worker 内部已隔离异常，不抛整体异常
                metadata_result = future.result()

        # 全部歌曲任务完成后统一写段落 Excel
        if PROJECT_SECTIONS in enabled and section_rows:
            try:
                with _EXCEL_LOCK:
                    out = write_sections_excel(section_rows, output_dir)
                collectors[PROJECT_SECTIONS].add_success(out)
                collectors[PROJECT_SECTIONS].add_notes([f"共 {len(section_rows)} 段"])
            except Exception as exc:
                collectors[PROJECT_SECTIONS].add_failure("", str(exc))

    projects = []
    for key in tasks:
        if key == PROJECT_METADATA:
            projects.append(metadata_result or collectors[key].result())
        else:
            projects.append(collectors[key].result())
    return DeliveryExportResult(projects=projects)
